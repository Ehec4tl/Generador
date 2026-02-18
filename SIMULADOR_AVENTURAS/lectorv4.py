import random
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------
# Modificadores extendidos con 9 niveles y probabilidades
MODIFICADORES = {
    1: -50,
    2: -35,
    3: -20,
    4: -10,
    5: 0,
    6: 10,
    7: 20,
    8: 35,
    9: 50
}

PESOS_BASE = [0.01, 0.02, 0.05, 0.10, 0.40, 0.10, 0.05, 0.02, 0.01]
PROBABILIDAD_MYB = 0.5  # 50% tiene MyB, 50% no
TOTAL_MAXIMO = 6000  # Límite para generación masiva

# --------------------------------------------------------
# CARGA JSON

def load_dictionary():
    with open("diccionario_personajesv4.json", "r", encoding="utf-8") as f:
        return json.load(f)

def load_classes():
    with open("clases.json", "r", encoding="utf-8") as f:
        return json.load(f)

# --------------------------------------------------------
# SISTEMA DE MODIFICADORES CON AJUSTE DE CLASE

def weighted_attribute(adjustment=None):
    valores = list(MODIFICADORES.keys())
    pesos = PESOS_BASE.copy()

    if adjustment:
        for i, lvl in enumerate(valores):
            pesos[i] = max(0, pesos[i] + adjustment.get(str(lvl), 0) * 0.001)

    return random.choices(valores, weights=pesos, k=1)[0]

# --------------------------------------------------------
# CLASE ALEATORIA

def assign_random_class(classes):
    pools = ["clases_simples", "clases_dobles", "clases_dobles_puras"]
    pool = random.choice(pools)
    key = random.choice(list(classes[pool].keys()))
    return key, classes[pool][key]

# --------------------------------------------------------
def generate_code(first_digit, classes):
    d1 = first_digit
    d2 = random.randint(1, 3)
    d3 = random.randint(1, 8)
    d4 = random.randint(1, 3)

    class_key, class_data = assign_random_class(classes)
    perfil = class_data["perfil"]
    atributos_clase = class_data["atributos"]
    ajuste = classes["perfiles_modificador"][perfil]["ajuste_probabilidad"]

    attrs = {}

    for key in ["F","RM","RF","A","I","M","E","C"]:
        if key in atributos_clase:
            attrs[key] = weighted_attribute(ajuste)
        else:
            attrs[key] = weighted_attribute()

    code = f"{d1}{d2}{d3}{d4}"
    code += "".join([f"{k}{v}" for k,v in attrs.items()])

    return code, attrs, d1, d2, d3, d4, class_key, class_data["nombre"]

# --------------------------------------------------------
def translate_character(dic, d1, d2, d3, d4):
    rp = str(d1)
    rs = str(d2)
    myb = str(d3)
    var = str(d4)

    raza = dic["razas_principales"].get(rp, "Desconocido")

    # Bonus de subraza
    subraza_data = dic["subrazas"].get(rp, {}).get(rs, {})

    # Verificamos si subraza_data es un diccionario (como debería ser)
    if isinstance(subraza_data, dict):
        subraza = subraza_data.get("nombre", "Desconocido")
        bonus = subraza_data.get("bonus", {})
    else:
        # Fallback por si acaso hay algún dato antiguo
        subraza = str(subraza_data) if subraza_data else "Desconocido"
        bonus = {}

    # ===== MyB 50/50 =====
    if random.random() < PROBABILIDAD_MYB:
        myb_data = dic["myb"].get(myb, {})
        myb_nombre = myb_data.get("nombre", "Desconocido")
        myb_variante = myb_data.get("variantes", {}).get(var, "Desconocido")
    else:
        myb_nombre = "Ninguna"
        myb_variante = "Sin marca"
    # =====================

    return raza, subraza, bonus, myb_nombre, myb_variante

# --------------------------------------------------------
def apply_base_mod(dic, raza, bonus_subraza, attrs):
    valores_base = dic["valores_base"].get(raza, {})
    abs_attrs = {}
    fuera_escala = {}

    for k, v in attrs.items():
        base = valores_base.get(k, 50)
        valor_abs = base + MODIFICADORES[v]
        valor_abs += bonus_subraza.get(k, 0)

        # SIN LÍMITES - el valor puede ser cualquier número
        abs_attrs[k] = valor_abs
        fuera_escala[k] = False

    suma_total = sum(abs_attrs.values())
    return abs_attrs, fuera_escala, suma_total

# --------------------------------------------------------
def classify_character(abs_attrs, fuera_escala, suma_total):
    valores_extremos = sum(1 for v in abs_attrs.values() if v >= 120 or v <= -20)

    # ESCALA DE HÉROES
    if suma_total >= 530:
        return "Héroe Legendario"
    elif suma_total >= 490:
        return "Héroe Mayor"
    elif suma_total >= 450:
        return "Héroe"

    # ABERRANTES
    elif valores_extremos >= 2:
        return "Aberrante"

    # NORMALES
    else:
        return "Normal"

# --------------------------------------------------------
def generate_characters():
    dic = load_dictionary()
    classes = load_classes()
    results = []

    print(f"\n🎲 GENERANDO PERSONAJES (máximo {TOTAL_MAXIMO})")
    print("="*50)
    print("Ingresa cuántos personajes quieres de cada tipo (1 al 8):")

    total = 0
    total_personajes = 0
    
    # CAMBIO IMPORTANTE: Ahora iteramos del 1 al 8 para las 8 razas
    for i in range(1, 9):  # Antes era range(1,7), ahora es range(1,9)
        while True:
            try:
                count = int(input(f"Tipo {i} (Raza {i}): "))
                if count < 0:
                    print("  ⚠️ El número no puede ser negativo")
                    continue
                if total + count > TOTAL_MAXIMO:
                    print(f"  ⚠️ Máximo disponible: {TOTAL_MAXIMO - total}")
                    continue
                break
            except ValueError:
                print("  ⚠️ Ingresa un número válido")

        total += count

        for j in range(count):
            # Mostrar progreso cada 500 personajes
            if total_personajes % 500 == 0 and total_personajes > 0:
                print(f"  Progreso: {total_personajes}/{TOTAL_MAXIMO} personajes generados...")

            code, attrs, d1, d2, d3, d4, class_key, class_name = generate_code(i, classes)
            raza, subraza, bonus, myb, variante = translate_character(dic, d1, d2, d3, d4)
            abs_attrs, fuera_escala, suma_total = apply_base_mod(dic, raza, bonus, attrs)
            tipo = classify_character(abs_attrs, fuera_escala, suma_total)

            results.append([
                code,
                raza,
                subraza,
                myb,
                variante,
                class_name,
                class_key,
                attrs["F"], abs_attrs["F"],
                attrs["RM"], abs_attrs["RM"],
                attrs["RF"], abs_attrs["RF"],
                attrs["A"], abs_attrs["A"],
                attrs["I"], abs_attrs["I"],
                attrs["M"], abs_attrs["M"],
                attrs["E"], abs_attrs["E"],
                attrs["C"], abs_attrs["C"],
                suma_total,
                tipo
            ])

            total_personajes += 1

    print(f"  ✅ Generación completada: {total_personajes} personajes")
    return results

# --------------------------------------------------------
# FUNCIONES DE FORMATO EXCEL
# --------------------------------------------------------

def formatear_hoja_principal(ws):
    # Negritas en encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Congelar fila superior
    ws.freeze_panes = "A2"

    # Bordes
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    # Centrar atributos
    columnas_centradas = ["D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X"]
    for col in columnas_centradas:
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal="center")

    # Autoajustar ancho
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

def colorear_por_tipo(ws):
    fill_legendario = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Dorado
    fill_mayor = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")      # Plateado
    fill_heroe = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")       # Bronce
    fill_aberrante = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")   # Rojo claro
    fill_normal = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")      # Gris claro

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tipo = row[24].value
        
        if tipo == "Héroe Legendario":
            fill = fill_legendario
        elif tipo == "Héroe Mayor":
            fill = fill_mayor
        elif tipo == "Héroe":
            fill = fill_heroe
        elif tipo == "Aberrante":
            fill = fill_aberrante
        else:
            fill = fill_normal

        for cell in row:
            cell.fill = fill

# --------------------------------------------------------
# HOJA DE ESTADÍSTICAS GENERALES
# --------------------------------------------------------

def crear_hoja_estadisticas(wb, data):
    """Crea una hoja con estadísticas detalladas de la generación"""
    
    ws_stats = wb.create_sheet("Estadísticas")
    
    # Estilos
    titulo_font = Font(bold=True, size=14)
    subtitulo_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    
    # =====================================================
    # 1. RESUMEN GENERAL
    # =====================================================
    ws_stats["A1"] = "📊 ESTADÍSTICAS DE GENERACIÓN"
    ws_stats["A1"].font = titulo_font
    ws_stats.merge_cells("A1:C1")
    ws_stats["A1"].alignment = Alignment(horizontal="center")
    
    ws_stats["A3"] = "TOTAL PERSONAJES:"
    ws_stats["B3"] = len(data)
    ws_stats["A3"].font = subtitulo_font
    ws_stats["B3"].font = subtitulo_font
    
    # =====================================================
    # 2. DISTRIBUCIÓN POR TIPO
    # =====================================================
    ws_stats["A5"] = "📌 DISTRIBUCIÓN POR TIPO"
    ws_stats["A5"].font = subtitulo_font
    
    # Headers
    ws_stats["A6"] = "TIPO"
    ws_stats["B6"] = "CANTIDAD"
    ws_stats["C6"] = "PORCENTAJE"
    for col in ["A6","B6","C6"]:
        ws_stats[col].font = header_font
        ws_stats[col].alignment = Alignment(horizontal="center")
    
    # Calcular tipos
    tipos = {}
    for p in data:
        tipo = p[24]
        tipos[tipo] = tipos.get(tipo, 0) + 1
    
    # Llenar datos
    fila = 7
    for tipo, count in sorted(tipos.items(), key=lambda x: x[1], reverse=True):
        ws_stats[f"A{fila}"] = tipo
        ws_stats[f"B{fila}"] = count
        ws_stats[f"C{fila}"] = f"{(count/len(data)*100):.1f}%"
        ws_stats[f"B{fila}"].alignment = Alignment(horizontal="center")
        ws_stats[f"C{fila}"].alignment = Alignment(horizontal="center")
        fila += 1
    
    # =====================================================
    # 3. TOP 10 CLASES
    # =====================================================
    ws_stats["E5"] = "🏆 TOP 10 CLASES MÁS COMUNES"
    ws_stats["E5"].font = subtitulo_font
    
    ws_stats["E6"] = "CLASE"
    ws_stats["F6"] = "CANTIDAD"
    ws_stats["G6"] = "%"
    for col in ["E6","F6","G6"]:
        ws_stats[col].font = header_font
        ws_stats[col].alignment = Alignment(horizontal="center")
    
    # Calcular clases
    clases = {}
    for p in data:
        clase = p[5]
        clases[clase] = clases.get(clase, 0) + 1
    
    fila = 7
    for clase, count in sorted(clases.items(), key=lambda x: x[1], reverse=True)[:10]:
        ws_stats[f"E{fila}"] = clase
        ws_stats[f"F{fila}"] = count
        ws_stats[f"G{fila}"] = f"{(count/len(data)*100):.1f}%"
        ws_stats[f"F{fila}"].alignment = Alignment(horizontal="center")
        ws_stats[f"G{fila}"].alignment = Alignment(horizontal="center")
        fila += 1
    
    # =====================================================
    # 4. TOP 10 RAZAS
    # =====================================================
    ws_stats["I5"] = "🌍 TOP 10 RAZAS MÁS COMUNES"
    ws_stats["I5"].font = subtitulo_font
    
    ws_stats["I6"] = "RAZA"
    ws_stats["J6"] = "CANTIDAD"
    ws_stats["K6"] = "%"
    for col in ["I6","J6","K6"]:
        ws_stats[col].font = header_font
        ws_stats[col].alignment = Alignment(horizontal="center")
    
    # Calcular razas
    razas = {}
    for p in data:
        raza = p[1]
        razas[raza] = razas.get(raza, 0) + 1
    
    fila = 7
    for raza, count in sorted(razas.items(), key=lambda x: x[1], reverse=True)[:10]:
        ws_stats[f"I{fila}"] = raza
        ws_stats[f"J{fila}"] = count
        ws_stats[f"K{fila}"] = f"{(count/len(data)*100):.1f}%"
        ws_stats[f"J{fila}"].alignment = Alignment(horizontal="center")
        ws_stats[f"K{fila}"].alignment = Alignment(horizontal="center")
        fila += 1
    
    # =====================================================
    # 5. ESTADÍSTICAS DE ATRIBUTOS
    # =====================================================
    ws_stats["A15"] = "📈 ESTADÍSTICAS DE ATRIBUTOS"
    ws_stats["A15"].font = subtitulo_font
    ws_stats.merge_cells("A15:C15")
    
    # Headers
    ws_stats["A16"] = "ATRIBUTO"
    ws_stats["B16"] = "PROMEDIO"
    ws_stats["C16"] = "MÍNIMO"
    ws_stats["D16"] = "MÁXIMO"
    ws_stats["E16"] = "≥90"
    ws_stats["F16"] = "≤10"
    
    for col in ["A16","B16","C16","D16","E16","F16"]:
        ws_stats[col].font = header_font
        ws_stats[col].alignment = Alignment(horizontal="center")
    
    # Calcular estadísticas por atributo
    atributos = ["F", "RM", "RF", "A", "I", "M", "E", "C"]
    indices_abs = [8, 10, 12, 14, 16, 18, 20, 22]
    
    fila = 17
    for attr, idx in zip(atributos, indices_abs):
        valores = [p[idx] for p in data]
        
        promedio = sum(valores) / len(valores)
        minimo = min(valores)
        maximo = max(valores)
        extremo_alto = sum(1 for v in valores if v >= 90)
        extremo_bajo = sum(1 for v in valores if v <= 10)
        
        ws_stats[f"A{fila}"] = attr
        ws_stats[f"B{fila}"] = round(promedio, 1)
        ws_stats[f"C{fila}"] = minimo
        ws_stats[f"D{fila}"] = maximo
        ws_stats[f"E{fila}"] = extremo_alto
        ws_stats[f"F{fila}"] = extremo_bajo
        
        for col in ["B","C","D","E","F"]:
            ws_stats[f"{col}{fila}"].alignment = Alignment(horizontal="center")
        
        fila += 1
    
    # Formato final
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for row in ws_stats.iter_rows(min_row=6, max_row=fila-1):
        for cell in row:
            if cell.value:
                cell.border = border
    
    for col in ["A","B","C","D","E","F","G","H","I","J","K"]:
        ws_stats.column_dimensions[col].width = 15

# --------------------------------------------------------
# HOJA DE ESTADÍSTICAS POR RAZA
# --------------------------------------------------------

def crear_estadisticas_por_raza(wb, data):
    """Crea una hoja con estadísticas detalladas por raza y subraza"""
    
    ws_razas = wb.create_sheet("Razas")
    
    # Estilos
    titulo_font = Font(bold=True, size=14)
    subtitulo_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    
    # Título
    ws_razas["A1"] = "🌍 ESTADÍSTICAS POR RAZA Y SUBRAZA"
    ws_razas["A1"].font = titulo_font
    ws_razas.merge_cells("A1:M1")
    ws_razas["A1"].alignment = Alignment(horizontal="center")
    
    # =====================================================
    # 1. RESUMEN POR RAZA PRINCIPAL
    # =====================================================
    ws_razas["A3"] = "📌 RESUMEN POR RAZA"
    ws_razas["A3"].font = subtitulo_font
    
    # Headers
    headers_raza = ["Raza", "Cantidad", "% Total", "Héroes", "% Héroes", 
                    "Aberrantes", "% Aberrantes", "Suma Promedio"]
    for i, header in enumerate(headers_raza):
        cell = ws_razas.cell(row=4, column=i+1, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Calcular datos por raza
    razas = {}
    for p in data:
        raza = p[1]
        if raza not in razas:
            razas[raza] = {
                "total": 0,
                "heroes": 0,
                "aberrantes": 0,
                "sumas": []
            }
        razas[raza]["total"] += 1
        razas[raza]["sumas"].append(p[23])
        
        if p[24] in ["Héroe", "Héroe Mayor", "Héroe Legendario"]:
            razas[raza]["heroes"] += 1
        elif p[24] == "Aberrante":
            razas[raza]["aberrantes"] += 1
    
    # Llenar datos de razas
    fila = 5
    for raza, stats in sorted(razas.items(), key=lambda x: x[1]["total"], reverse=True):
        total = stats["total"]
        heroes = stats["heroes"]
        aberrantes = stats["aberrantes"]
        suma_prom = sum(stats["sumas"]) / total if total > 0 else 0
        
        ws_razas[f"A{fila}"] = raza
        ws_razas[f"B{fila}"] = total
        ws_razas[f"C{fila}"] = f"{(total/len(data)*100):.1f}%"
        ws_razas[f"D{fila}"] = heroes
        ws_razas[f"E{fila}"] = f"{(heroes/total*100):.1f}%"
        ws_razas[f"F{fila}"] = aberrantes
        ws_razas[f"G{fila}"] = f"{(aberrantes/total*100):.1f}%"
        ws_razas[f"H{fila}"] = round(suma_prom, 1)
        
        for col in ["B","C","D","E","F","G","H"]:
            ws_razas[f"{col}{fila}"].alignment = Alignment(horizontal="center")
        
        fila += 1
    
    # =====================================================
    # 2. DETALLE POR SUBRAZA
    # =====================================================
    fila += 2
    ws_razas[f"A{fila}"] = "📋 DETALLE POR SUBRAZA"
    ws_razas[f"A{fila}"].font = subtitulo_font
    fila += 1
    
    # Headers subrazas
    headers_sub = ["Raza", "Subraza", "Cantidad", "% en Raza", "Héroes", "% Héroes", 
                   "Aberrantes", "% Aberrantes", "F", "RM", "RF", "A", "I", "M", "E", "C", "Suma"]
    
    for i, header in enumerate(headers_sub):
        cell = ws_razas.cell(row=fila, column=i+1, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    fila += 1
    
    # Agrupar por raza y subraza
    subrazas = {}
    for p in data:
        raza = p[1]
        subraza = p[2]
        key = f"{raza}|{subraza}"
        
        if key not in subrazas:
            subrazas[key] = {
                "raza": raza,
                "subraza": subraza,
                "total": 0,
                "heroes": 0,
                "aberrantes": 0,
                "F": [], "RM": [], "RF": [], "A": [], "I": [], "M": [], "E": [], "C": [],
                "sumas": []
            }
        
        stats = subrazas[key]
        stats["total"] += 1
        stats["sumas"].append(p[23])
        
        stats["F"].append(p[8])
        stats["RM"].append(p[10])
        stats["RF"].append(p[12])
        stats["A"].append(p[14])
        stats["I"].append(p[16])
        stats["M"].append(p[18])
        stats["E"].append(p[20])
        stats["C"].append(p[22])
        
        if p[24] in ["Héroe", "Héroe Mayor", "Héroe Legendario"]:
            stats["heroes"] += 1
        elif p[24] == "Aberrante":
            stats["aberrantes"] += 1
    
    # Llenar datos de subrazas
    for key, stats in sorted(subrazas.items()):
        total = stats["total"]
        
        ws_razas[f"A{fila}"] = stats["raza"]
        ws_razas[f"B{fila}"] = stats["subraza"]
        ws_razas[f"C{fila}"] = total
        
        total_raza = razas[stats["raza"]]["total"]
        ws_razas[f"D{fila}"] = f"{(total/total_raza*100):.1f}%"
        
        ws_razas[f"E{fila}"] = stats["heroes"]
        ws_razas[f"F{fila}"] = f"{(stats['heroes']/total*100):.1f}%"
        ws_razas[f"G{fila}"] = stats["aberrantes"]
        ws_razas[f"H{fila}"] = f"{(stats['aberrantes']/total*100):.1f}%"
        
        ws_razas[f"I{fila}"] = round(sum(stats["F"])/total, 1)
        ws_razas[f"J{fila}"] = round(sum(stats["RM"])/total, 1)
        ws_razas[f"K{fila}"] = round(sum(stats["RF"])/total, 1)
        ws_razas[f"L{fila}"] = round(sum(stats["A"])/total, 1)
        ws_razas[f"M{fila}"] = round(sum(stats["I"])/total, 1)
        ws_razas[f"N{fila}"] = round(sum(stats["M"])/total, 1)
        ws_razas[f"O{fila}"] = round(sum(stats["E"])/total, 1)
        ws_razas[f"P{fila}"] = round(sum(stats["C"])/total, 1)
        ws_razas[f"Q{fila}"] = round(sum(stats["sumas"])/total, 1)
        
        for col in range(1, 18):
            cell = ws_razas.cell(row=fila, column=col)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                top=Side(style="thin"), bottom=Side(style="thin"))
        
        fila += 1
    
    for col in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q"]:
        ws_razas.column_dimensions[col].width = 12



def crear_estadisticas_heroes(wb, data):
    """Crea una hoja con análisis detallado de los héroes"""
    
    ws_heroes = wb.create_sheet("Análisis de Héroes")
    
    # Estilos
    titulo_font = Font(bold=True, size=14)
    subtitulo_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    
    # Título
    ws_heroes["A1"] = "⚔️ ANÁLISIS DE HÉROES - PERSONAJES EXCEPCIONALES"
    ws_heroes["A1"].font = titulo_font
    ws_heroes.merge_cells("A1:M1")
    ws_heroes["A1"].alignment = Alignment(horizontal="center")
    
    # =====================================================
    # 1. RESUMEN GENERAL DE HÉROES
    # =====================================================
    ws_heroes["A3"] = "📌 RESUMEN DE HÉROES"
    ws_heroes["A3"].font = subtitulo_font
    
    # Filtrar solo héroes
    heroes = [p for p in data if p[24] in ["Héroe", "Héroe Mayor", "Héroe Legendario"]]
    total_heroes = len(heroes)
    
    if total_heroes == 0:
        ws_heroes["A5"] = "No se encontraron héroes en esta generación"
        return
    
    # Estadísticas básicas
    ws_heroes["A5"] = "Total de Héroes:"
    ws_heroes["B5"] = total_heroes
    ws_heroes["A6"] = "Porcentaje del total:"
    ws_heroes["B6"] = f"{(total_heroes/len(data)*100):.1f}%"
    
    # Sumas de héroes
    sumas_heroes = [p[23] for p in heroes]
    ws_heroes["A8"] = "Suma mínima:"
    ws_heroes["B8"] = min(sumas_heroes)
    ws_heroes["A9"] = "Suma máxima:"
    ws_heroes["B9"] = max(sumas_heroes)
    ws_heroes["A10"] = "Suma promedio:"
    ws_heroes["B10"] = round(sum(sumas_heroes)/total_heroes, 1)
    
    # =====================================================
    # 2. DISTRIBUCIÓN POR NIVEL DE HÉROE
    # =====================================================
    ws_heroes["D3"] = "📊 DISTRIBUCIÓN POR NIVEL"
    ws_heroes["D3"].font = subtitulo_font
    
    niveles = {}
    for p in heroes:
        nivel = p[24]
        niveles[nivel] = niveles.get(nivel, 0) + 1
    
    ws_heroes["D5"] = "NIVEL"
    ws_heroes["E5"] = "CANTIDAD"
    ws_heroes["F5"] = "% DE HÉROES"
    ws_heroes["G5"] = "% DEL TOTAL"
    
    for col in ["D5","E5","F5","G5"]:
        ws_heroes[col].font = header_font
        ws_heroes[col].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    fila = 6
    for nivel in ["Héroe Legendario", "Héroe Mayor", "Héroe"]:
        count = niveles.get(nivel, 0)
        ws_heroes[f"D{fila}"] = nivel
        ws_heroes[f"E{fila}"] = count
        ws_heroes[f"F{fila}"] = f"{(count/total_heroes*100):.1f}%"
        ws_heroes[f"G{fila}"] = f"{(count/len(data)*100):.1f}%"
        
        # Colorear según nivel
        if nivel == "Héroe Legendario":
            color = "FFD700"  # Dorado
        elif nivel == "Héroe Mayor":
            color = "C0C0C0"  # Plateado
        else:
            color = "CD7F32"  # Bronce
        
        for col in ["D","E","F","G"]:
            ws_heroes[f"{col}{fila}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        fila += 1
    
    # =====================================================
    # 3. TOP 10 HÉROES POR RAZA
    # =====================================================
    ws_heroes["I3"] = "🏆 TOP 10 RAZAS CON MÁS HÉROES"
    ws_heroes["I3"].font = subtitulo_font
    
    # Contar héroes por raza
    heroes_por_raza = {}
    for p in heroes:
        raza = p[1]
        heroes_por_raza[raza] = heroes_por_raza.get(raza, 0) + 1
    
    ws_heroes["I5"] = "RAZA"
    ws_heroes["J5"] = "HÉROES"
    ws_heroes["K5"] = "% DE HÉROES"
    ws_heroes["L5"] = "EFICIENCIA*"
    
    for col in ["I5","J5","K5","L5"]:
        ws_heroes[col].font = header_font
        ws_heroes[col].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Calcular totales por raza para eficiencia
    total_por_raza = {}
    for p in data:
        raza = p[1]
        total_por_raza[raza] = total_por_raza.get(raza, 0) + 1
    
    fila = 6
    for raza, count in sorted(heroes_por_raza.items(), key=lambda x: x[1], reverse=True)[:10]:
        total_raza = total_por_raza.get(raza, 1)
        eficiencia = (count / total_raza) * 100
        
        ws_heroes[f"I{fila}"] = raza
        ws_heroes[f"J{fila}"] = count
        ws_heroes[f"K{fila}"] = f"{(count/total_heroes*100):.1f}%"
        ws_heroes[f"L{fila}"] = f"{eficiencia:.1f}%"
        
        # Colorear según eficiencia
        if eficiencia >= 20:
            ws_heroes[f"L{fila}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde
        elif eficiencia >= 15:
            ws_heroes[f"L{fila}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarillo
        
        fila += 1
    
    ws_heroes[f"L4"] = "*% de personajes de esa raza que son héroes"
    ws_heroes[f"L4"].font = Font(italic=True, size=9)
    
    # =====================================================
    # 4. TOP 10 CLASES DE HÉROES
    # =====================================================
    fila_actual = fila + 2
    ws_heroes[f"I{fila_actual}"] = "⚔️ TOP 10 CLASES DE HÉROES"
    ws_heroes[f"I{fila_actual}"].font = subtitulo_font
    fila_actual += 2
    
    # Contar héroes por clase
    heroes_por_clase = {}
    for p in heroes:
        clase = p[5]
        heroes_por_clase[clase] = heroes_por_clase.get(clase, 0) + 1
    
    ws_heroes[f"I{fila_actual}"] = "CLASE"
    ws_heroes[f"J{fila_actual}"] = "HÉROES"
    ws_heroes[f"K{fila_actual}"] = "% DE HÉROES"
    
    for col in [f"I{fila_actual}", f"J{fila_actual}", f"K{fila_actual}"]:
        ws_heroes[col].font = header_font
        ws_heroes[col].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    fila_actual += 1
    for clase, count in sorted(heroes_por_clase.items(), key=lambda x: x[1], reverse=True)[:10]:
        ws_heroes[f"I{fila_actual}"] = clase
        ws_heroes[f"J{fila_actual}"] = count
        ws_heroes[f"K{fila_actual}"] = f"{(count/total_heroes*100):.1f}%"
        fila_actual += 1
    
    # =====================================================
    # 5. PERFIL DE ATRIBUTOS DE HÉROES
    # =====================================================
    fila_actual = 6
    col_inicio = 15  # Columna O
    
    ws_heroes[f"{get_column_letter(col_inicio)}3"] = "📈 PERFIL DE ATRIBUTOS DE HÉROES"
    ws_heroes[f"{get_column_letter(col_inicio)}3"].font = subtitulo_font
    
    # Headers
    headers_atributos = ["Atributo", "Promedio Héroes", "Promedio General", "Diferencia", "Mínimo", "Máximo"]
    for i, header in enumerate(headers_atributos):
        col_letra = get_column_letter(col_inicio + i)
        cell = ws_heroes[f"{col_letra}5"]
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Calcular promedios de atributos para héroes
    atributos = ["F", "RM", "RF", "A", "I", "M", "E", "C"]
    indices_abs = [8, 10, 12, 14, 16, 18, 20, 22]
    
    fila_actual = 6
    for attr, idx in zip(atributos, indices_abs):
        # Valores de héroes
        valores_heroes = [p[idx] for p in heroes]
        prom_heroes = sum(valores_heroes) / len(valores_heroes) if valores_heroes else 0
        min_heroes = min(valores_heroes) if valores_heroes else 0
        max_heroes = max(valores_heroes) if valores_heroes else 0
        
        # Valores generales
        valores_gral = [p[idx] for p in data]
        prom_gral = sum(valores_gral) / len(valores_gral)
        
        diferencia = prom_heroes - prom_gral
        
        col_letra = get_column_letter(col_inicio)
        ws_heroes[f"{col_letra}{fila_actual}"] = attr
        
        col_letra = get_column_letter(col_inicio + 1)
        ws_heroes[f"{col_letra}{fila_actual}"] = round(prom_heroes, 1)
        
        col_letra = get_column_letter(col_inicio + 2)
        ws_heroes[f"{col_letra}{fila_actual}"] = round(prom_gral, 1)
        
        col_letra = get_column_letter(col_inicio + 3)
        ws_heroes[f"{col_letra}{fila_actual}"] = round(diferencia, 1)
        # Colorear diferencia positiva/negativa
        if diferencia > 5:
            ws_heroes[f"{col_letra}{fila_actual}"].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif diferencia < -5:
            ws_heroes[f"{col_letra}{fila_actual}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        col_letra = get_column_letter(col_inicio + 4)
        ws_heroes[f"{col_letra}{fila_actual}"] = min_heroes
        
        col_letra = get_column_letter(col_inicio + 5)
        ws_heroes[f"{col_letra}{fila_actual}"] = max_heroes
        
        # Centrar todos los valores
        for i in range(6):
            col_letra = get_column_letter(col_inicio + i)
            ws_heroes[f"{col_letra}{fila_actual}"].alignment = Alignment(horizontal="center")
        
        fila_actual += 1
    
    # =====================================================
    # 6. HÉROES DESTACADOS (TOP 10)
    # =====================================================
    fila_actual += 2
    ws_heroes[f"A{fila_actual}"] = "🏅 TOP 10 HÉROES MÁS PODEROSOS"
    ws_heroes[f"A{fila_actual}"].font = subtitulo_font
    fila_actual += 1
    
    # Headers
    headers_top = ["#", "Raza", "Subraza", "Clase", "MyB", "Suma", "Tipo", "Atributo Máx"]
    for i, header in enumerate(headers_top):
        col_letra = get_column_letter(i + 1)
        cell = ws_heroes[f"{col_letra}{fila_actual}"]
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Ordenar héroes por suma total
    heroes_ordenados = sorted(heroes, key=lambda x: x[23], reverse=True)[:10]
    
    fila_actual += 1
    for i, heroe in enumerate(heroes_ordenados, 1):
        # Encontrar atributo máximo
        valores_atrs = [heroe[8], heroe[10], heroe[12], heroe[14], 
                       heroe[16], heroe[18], heroe[20], heroe[22]]
        atrs_nombres = ["F", "RM", "RF", "A", "I", "M", "E", "C"]
        max_valor = max(valores_atrs)
        max_attr = atrs_nombres[valores_atrs.index(max_valor)]
        
        ws_heroes[f"A{fila_actual}"] = i
        ws_heroes[f"B{fila_actual}"] = heroe[1]  # Raza
        ws_heroes[f"C{fila_actual}"] = heroe[2]  # Subraza
        ws_heroes[f"D{fila_actual}"] = heroe[5]  # Clase
        ws_heroes[f"E{fila_actual}"] = heroe[3]  # MyB
        ws_heroes[f"F{fila_actual}"] = heroe[23]  # Suma
        ws_heroes[f"G{fila_actual}"] = heroe[24]  # Tipo
        ws_heroes[f"H{fila_actual}"] = f"{max_attr}: {max_valor}"
        
        # Colorear según tipo
        if heroe[24] == "Héroe Legendario":
            color = "FFD700"
        elif heroe[24] == "Héroe Mayor":
            color = "C0C0C0"
        else:
            color = "CD7F32"
        
        for col in range(1, 9):
            col_letra = get_column_letter(col)
            ws_heroes[f"{col_letra}{fila_actual}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        fila_actual += 1
    
    # =====================================================
    # 7. FORMATO FINAL
    # =====================================================
    # Autoajustar anchos
    for col in range(1, 30):
        col_letra = get_column_letter(col)
        ws_heroes.column_dimensions[col_letra].width = 15
    
    ws_heroes.column_dimensions["A"].width = 8
    ws_heroes.column_dimensions["B"].width = 15
    ws_heroes.column_dimensions["C"].width = 20
    ws_heroes.column_dimensions["D"].width = 20
    ws_heroes.column_dimensions["E"].width = 15

# --------------------------------------------------------
# FUNCIÓN PRINCIPAL DE GUARDADO
# --------------------------------------------------------

def save_to_excel(data):
    print("\n💾 Guardando en Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Personajes"

    headers = [
        "Código", "Raza", "Subraza", "MyB", "Variante",
        "Clase", "Clave_Clase",
        "F_val", "F_abs",
        "RM_val", "RM_abs",
        "RF_val", "RF_abs",
        "A_val", "A_abs",
        "I_val", "I_abs",
        "M_val", "M_abs",
        "E_val", "E_abs",
        "C_val", "C_abs",
        "Suma_Total",
        "Tipo"
    ]

    ws.append(headers)

    # Escritura por lotes para mejor feedback
    batch_size = 1000
    total_filas = len(data)
    
    for i in range(0, total_filas, batch_size):
        batch = data[i:i+batch_size]
        for row in batch:
            ws.append(row)
        print(f"  Excel: {min(i+batch_size, total_filas)}/{total_filas} filas escritas...")

    print("  Aplicando formato a la hoja principal...")
    formatear_hoja_principal(ws)
    colorear_por_tipo(ws)
    
    print("  Generando hoja de estadísticas generales...")
    crear_hoja_estadisticas(wb, data)
    
    print("  Generando hoja de estadísticas por raza...")
    crear_estadisticas_por_raza(wb, data)
    
    print("  Generando análisis detallado de héroes...")
    crear_estadisticas_heroes(wb, data)

    # Guardar archivo
    wb.save("personajes_generadosv4.xlsx")
    
    print("\n" + "="*50)
    print("✅ ARCHIVO GUARDADO EXITOSAMENTE")
    print("="*50)
    print(f"📁 Archivo: personajes_generadosv4.xlsx")
    print(f"📊 Total personajes: {len(data)}")
    
    # Mostrar resumen rápido de tipos
    tipos = {}
    for p in data:
        tipos[p[24]] = tipos.get(p[24], 0) + 1
    
    print("\n📊 RESUMEN POR TIPO:")
    for tipo, count in sorted(tipos.items(), key=lambda x: x[1], reverse=True):
        print(f"  • {tipo}: {count} ({count/len(data)*100:.1f}%)")
    
    print("\n✨ Proceso completado!")

# --------------------------------------------------------
if __name__ == "__main__":
    print("🎲 GENERADOR DE PERSONAJES - VERSIÓN 4.0")
    print("="*50)
    print("🔧 Características:")
    print("  • 6000 personajes máximo")
    print("  • 8 razas disponibles (Humanos a Medianos)")
    print("  • Bonus de subraza")
    print("  • MyB 50/50 narrativo")
    print("  • Clasificación en 5 niveles")
    print("  • Estadísticas detalladas")
    print("="*50)
    
    personajes = generate_characters()
    save_to_excel(personajes)