"""
gestor_resultados.py - Exportación de resultados y generación de estadísticas.
"""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modulos.gestor_personajes import IndicesPersonaje


class EstilosExcel:
    COLORES_TIPO = {
        "Héroe Legendario": "FFD700",
        "Héroe Mayor": "C0C0C0",
        "Héroe": "CD7F32",
        "Aberrante": "FF9999",
        "Normal": "F0F0F0"
    }
    
    @classmethod
    def header_style(cls):
        return Font(bold=True)
    
    @classmethod
    def border_style(cls):
        thin = Side(style="thin")
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    
    @classmethod
    def get_fill_por_tipo(cls, tipo):
        color = cls.COLORES_TIPO.get(tipo, "FFFFFF")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")


class GestorResultados:
    def __init__(self):
        self.personajes_iniciales = None
        self.personajes_finales = None
        self.historial_eventos = []
        self.stats = {}
        self.dir_salida = "salidas"
        self.vivos = []
        self.muertos = []
        self._crear_directorio_salida()
    
    def _crear_directorio_salida(self):
        if not os.path.exists(self.dir_salida):
            os.makedirs(self.dir_salida)
    
    def cargar_datos(self, personajes_iniciales, personajes_finales, historial_eventos):
        """Carga todos los datos de la simulación."""
        self.personajes_iniciales = personajes_iniciales
        self.personajes_finales = personajes_finales  # Estos son TODOS (vivos y muertos)
        self.historial_eventos = historial_eventos
        
        # Contar vivos y muertos
        self.vivos = [p for p in personajes_finales if p[IndicesPersonaje.VIVO]]
        self.muertos = [p for p in personajes_finales if not p[IndicesPersonaje.VIVO]]
        
        self.calcular_estadisticas()
    
    # ========================================================================
    # MÉTODOS DE ESTADÍSTICAS
    # ========================================================================
    
    def _estadisticas_personajes(self, personajes):
        """Calcula estadísticas de una lista de personajes."""
        stats = {}
        
        if not personajes:
            return stats
        
        # Suma de atributos
        sumas = []
        for p in personajes:
            val = p[IndicesPersonaje.SUMA_TOTAL]
            if val is not None:
                sumas.append(val)
        
        stats["suma_promedio"] = sum(sumas) / len(sumas) if sumas else 0
        stats["suma_min"] = min(sumas) if sumas else 0
        stats["suma_max"] = max(sumas) if sumas else 0
        
        # Distribución por tipo
        tipos = {}
        for p in personajes:
            tipo = p[IndicesPersonaje.TIPO]
            if tipo:
                tipos[tipo] = tipos.get(tipo, 0) + 1
        stats["distribucion_tipos"] = tipos
        
        # Distribución por raza
        razas = {}
        for p in personajes:
            raza = p[IndicesPersonaje.RAZA]
            if raza:
                razas[raza] = razas.get(raza, 0) + 1
        stats["distribucion_razas"] = razas
        
        # Distribución por clase
        clases = {}
        for p in personajes:
            clase = p[IndicesPersonaje.CLASE]
            if clase:
                clases[clase] = clases.get(clase, 0) + 1
        stats["distribucion_clases"] = clases
        
        # Promedio de cada atributo
        stats["promedio_atributos"] = {}
        for attr_name, idx in [
            ("F", IndicesPersonaje.F_ABS),
            ("RM", IndicesPersonaje.RM_ABS),
            ("RF", IndicesPersonaje.RF_ABS),
            ("A", IndicesPersonaje.A_ABS),
            ("I", IndicesPersonaje.I_ABS),
            ("M", IndicesPersonaje.M_ABS),
            ("E", IndicesPersonaje.E_ABS),
            ("C", IndicesPersonaje.C_ABS)
        ]:
            valores = []
            for p in personajes:
                val = p[idx]
                if val is not None:
                    valores.append(val)
            stats["promedio_atributos"][attr_name] = sum(valores) / len(valores) if valores else 0
        
        # Eventos vividos
        eventos_vividos = []
        for p in personajes:
            val = p[IndicesPersonaje.EVENTOS_VIDA]
            if val is None:
                eventos_vividos.append(0)
            else:
                eventos_vividos.append(val)
        
        stats["eventos_por_personaje"] = sum(eventos_vividos) / len(eventos_vividos) if eventos_vividos else 0
        
        return stats
    
    def _estadisticas_eventos(self):
        """Calcula estadísticas de los eventos ocurridos."""
        stats = {}
        
        if not self.historial_eventos:
            return stats
        
        # Tipos de evento
        tipos = {}
        for e in self.historial_eventos:
            tipo = e.get("tipo", "desconocido")
            tipos[tipo] = tipos.get(tipo, 0) + 1
        stats["eventos_por_tipo"] = tipos
        
        # Resultados
        resultados = []
        for evento in self.historial_eventos:
            for p in evento.get("participantes", []):
                if "resultado" in p:
                    resultados.append(p["resultado"])
        
        stats_resultados = {}
        for r in resultados:
            stats_resultados[r] = stats_resultados.get(r, 0) + 1
        stats["resultados"] = stats_resultados
        
        # Muertes
        muertes = 0
        for e in self.historial_eventos:
            for p in e.get("participantes", []):
                if p.get("murio", False):
                    muertes += 1
        stats["muertes_en_eventos"] = muertes
        
        return stats
    
    def _estadisticas_caracteristicas(self):
        """Analiza las características (bendiciones/maldiciones) adquiridas."""
        stats = {}
        
        if not self.personajes_finales:
            return stats
        
        bendiciones = 0
        maldiciones = 0
        caracteristicas_por_personaje = []
        
        for p in self.personajes_finales:
            if len(p) > IndicesPersonaje.RESERVADO1:
                chars = p[IndicesPersonaje.RESERVADO1]
                
                # Diferentes formatos posibles
                if chars is None:
                    continue
                elif isinstance(chars, int):
                    # Es un contador, no una lista
                    caracteristicas_por_personaje.append(chars)
                    # No podemos saber si son bendiciones o maldiciones
                elif isinstance(chars, list):
                    # Es una lista de objetos/diccionarios
                    caracteristicas_por_personaje.append(len(chars))
                    for c in chars:
                        if c is None:
                            continue
                        elif isinstance(c, dict):
                            if c.get('es_bendicion', False):
                                bendiciones += 1
                            else:
                                maldiciones += 1
                        elif hasattr(c, 'es_bendicion'):
                            if c.es_bendicion:
                                bendiciones += 1
                            else:
                                maldiciones += 1
                else:
                    # Otro tipo, lo ignoramos
                    continue
        
        stats["total_bendiciones"] = bendiciones
        stats["total_maldiciones"] = maldiciones
        if caracteristicas_por_personaje:
            stats["promedio_caracteristicas"] = sum(caracteristicas_por_personaje) / len(caracteristicas_por_personaje)
        else:
            stats["promedio_caracteristicas"] = 0
        
        return stats
    
    def calcular_estadisticas(self):
        """Calcula todas las estadísticas de la simulación."""
        stats = {}
        
        # Estadísticas básicas (CORREGIDO)
        stats["total_inicial"] = len(self.personajes_iniciales) if self.personajes_iniciales else 0
        stats["vivos"] = len(self.vivos) if hasattr(self, 'vivos') else 0
        stats["muertos"] = len(self.muertos) if hasattr(self, 'muertos') else 0
        stats["total_final"] = stats["vivos"] + stats["muertos"]
        
        if stats["total_inicial"] > 0:
            stats["supervivencia"] = (stats["vivos"] / stats["total_inicial"]) * 100
        else:
            stats["supervivencia"] = 0
        
        stats["total_eventos"] = len(self.historial_eventos)
        
        # Estadísticas de personajes vivos (para promedios, etc.)
        if hasattr(self, 'vivos') and self.vivos:
            stats.update(self._estadisticas_personajes(self.vivos))
        
        # Estadísticas de eventos
        if self.historial_eventos:
            stats.update(self._estadisticas_eventos())
        
        # Estadísticas de características (de todos, no solo vivos)
        if self.personajes_finales:
            stats.update(self._estadisticas_caracteristicas())
        
        self.stats = stats
        return stats
    
    # ========================================================================
    # EXPORTACIÓN A EXCEL
    # ========================================================================
    
    def exportar_a_excel(self, nombre_base="simulacion", incluir_historial=True):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.dir_salida, f"{nombre_base}_{timestamp}.xlsx")
        
        wb = Workbook()
        
        if self.personajes_iniciales:
            self._crear_hoja_personajes(wb, "Iniciales", self.personajes_iniciales)
        
        if self.personajes_finales:
            self._crear_hoja_personajes(wb, "Finales", self.personajes_finales)
        
        if self.personajes_iniciales and self.personajes_finales:
            self._crear_hoja_comparativa(wb)
        
        if incluir_historial and self.historial_eventos:
            self._crear_hoja_eventos(wb)
        
        if self.stats:
            self._crear_hoja_estadisticas(wb)
        
        wb.save(filename)
        print(f"\n✅ Resultados guardados en: {filename}")
        
        return filename
    
    def _crear_hoja_personajes(self, wb, titulo, personajes):
        if not personajes:
            return
        
        ws = wb.create_sheet(titulo)
        
        headers = [
            "Código", "Raza", "Subraza", "MyB", "Variante",
            "Clase", "Clave_Clase",
            "F_val", "F_abs", "RM_val", "RM_abs", "RF_val", "RF_abs",
            "A_val", "A_abs", "I_val", "I_abs", "M_val", "M_abs",
            "E_val", "E_abs", "C_val", "C_abs",
            "Suma_Total", "Tipo",
            "Eventos", "Victorias", "Derrotas", "Vivo"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = EstilosExcel.header_style()
            cell.alignment = Alignment(horizontal="center")
        
        for row, p in enumerate(personajes, 2):
            while len(p) <= IndicesPersonaje.VIVO:
                p.append(None)
            
            for col, valor in enumerate(p[:len(headers)], 1):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.border = EstilosExcel.border_style()
                
                if col == 25:  # Columna Tipo
                    cell.fill = EstilosExcel.get_fill_por_tipo(valor)
        
        self._autoajustar_columnas(ws)
    
    def _crear_hoja_comparativa(self, wb):
        ws = wb.create_sheet("Comparativa")
        
        ws["A1"] = "📊 COMPARATIVA INICIALES VS FINALES"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:E1")
        
        headers = ["Métrica", "Inicial", "Final", "Diferencia", "Cambio %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = EstilosExcel.header_style()
        
        stats_inicial = self._estadisticas_personajes(self.personajes_iniciales)
        stats_final = self._estadisticas_personajes(self.personajes_finales)
        
        fila = 4
        metricas = [
            ("Total personajes", 
             len(self.personajes_iniciales), 
             len(self.personajes_finales)),
            ("Suma promedio", 
             stats_inicial.get("suma_promedio", 0), 
             stats_final.get("suma_promedio", 0)),
            ("Eventos por persona", 
             0, 
             stats_final.get("eventos_por_personaje", 0)),
        ]
        
        for nombre, inicial, final in metricas:
            ws.cell(row=fila, column=1, value=nombre)
            ws.cell(row=fila, column=2, value=round(inicial, 1))
            ws.cell(row=fila, column=3, value=round(final, 1))
            
            diferencia = final - inicial
            ws.cell(row=fila, column=4, value=round(diferencia, 1))
            
            if inicial != 0:
                cambio = (diferencia / inicial) * 100
                ws.cell(row=fila, column=5, value=f"{cambio:+.1f}%")
            
            if diferencia > 0:
                for col in [2,3,4,5]:
                    ws.cell(row=fila, column=col).fill = PatternFill(
                        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                    )
            elif diferencia < 0:
                for col in [2,3,4,5]:
                    ws.cell(row=fila, column=col).fill = PatternFill(
                        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                    )
            
            fila += 1
        
        self._autoajustar_columnas(ws)
    
    def _crear_hoja_eventos(self, wb):
        ws = wb.create_sheet("Eventos")
        
        headers = ["#", "Tipo", "Evento", "Participantes", "Resultados", "Muertes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = EstilosExcel.header_style()
        
        fila = 2
        for i, evento in enumerate(self.historial_eventos, 1):
            ws.cell(row=fila, column=1, value=i)
            ws.cell(row=fila, column=2, value=evento.get("tipo", ""))
            ws.cell(row=fila, column=3, value=evento.get("nombre", ""))
            
            participantes = len(evento.get("participantes", []))
            ws.cell(row=fila, column=4, value=participantes)
            
            resultados = [p.get("resultado", "") for p in evento.get("participantes", [])]
            ws.cell(row=fila, column=5, value=", ".join(resultados))
            
            muertes = sum(1 for p in evento.get("participantes", []) if p.get("murio", False))
            ws.cell(row=fila, column=6, value=muertes)
            
            fila += 1
        
        self._autoajustar_columnas(ws)
    
    def _crear_hoja_estadisticas(self, wb):
        ws = wb.create_sheet("Estadísticas")
        
        fila = 1
        ws["A1"] = "📈 ESTADÍSTICAS DE SIMULACIÓN"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:C1")
        fila = 3
        
        ws.cell(row=fila, column=1, value="📊 BÁSICAS").font = Font(bold=True)
        fila += 1
        
        basicas = [
            ("Personajes iniciales", self.stats.get("total_inicial", 0)),
            ("Personajes finales", self.stats.get("total_final", 0)),
            ("Muertos", self.stats.get("muertos", 0)),
            ("Supervivencia", f"{self.stats.get('supervivencia', 0):.1f}%"),
            ("Total eventos", self.stats.get("total_eventos", 0)),
        ]
        
        for nombre, valor in basicas:
            ws.cell(row=fila, column=1, value=nombre)
            ws.cell(row=fila, column=2, value=valor)
            fila += 1
        
        fila += 1
        ws.cell(row=fila, column=1, value="⚔️ ATRIBUTOS").font = Font(bold=True)
        fila += 1
        
        promedios = self.stats.get("promedio_atributos", {})
        for attr in ["F", "RM", "RF", "A", "I", "M", "E", "C"]:
            ws.cell(row=fila, column=1, value=f"{attr} promedio")
            ws.cell(row=fila, column=2, value=round(promedios.get(attr, 0), 1))
            fila += 1
        
        fila += 1
        ws.cell(row=fila, column=1, value="🏷️ DISTRIBUCIÓN POR TIPO").font = Font(bold=True)
        fila += 1
        
        tipos = self.stats.get("distribucion_tipos", {})
        total = self.stats.get("total_final", 1)
        for tipo, count in tipos.items():
            ws.cell(row=fila, column=1, value=tipo)
            ws.cell(row=fila, column=2, value=count)
            ws.cell(row=fila, column=3, value=f"{count/total*100:.1f}%")
            fila += 1
        
        self._autoajustar_columnas(ws)
    
    def _autoajustar_columnas(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
    
    def mostrar_resumen(self):
        """Muestra un resumen de la simulación en consola."""
        if not self.stats:
            self.calcular_estadisticas()
        
        print("\n" + "="*60)
        print("📊 RESUMEN DE SIMULACIÓN")
        print("="*60)
        
        print(f"\n📌 DATOS BÁSICOS:")
        print(f"  • Personajes iniciales: {self.stats.get('total_inicial', 0)}")
        print(f"  • Personajes vivos: {self.stats.get('vivos', 0)}")
        print(f"  • Personajes muertos: {self.stats.get('muertos', 0)}")
        print(f"  • Supervivencia: {self.stats.get('supervivencia', 0):.1f}%")
        print(f"  • Eventos totales: {self.stats.get('total_eventos', 0)}")
        
        print(f"\n⚔️ ATRIBUTOS (promedios finales):")
        promedios = self.stats.get("promedio_atributos", {})
        for attr in ["F", "RM", "RF", "A", "I", "M", "E", "C"]:
            print(f"  • {attr}: {promedios.get(attr, 0):.1f}")
        
        print(f"\n🏷️ TIPOS DE PERSONAJE:")
        tipos = self.stats.get("distribucion_tipos", {})
        total = self.stats.get("vivos", 1)
        for tipo, count in sorted(tipos.items(), key=lambda x: x[1], reverse=True):
            print(f"  • {tipo}: {count} ({count/total*100:.1f}%)")
        
        print(f"\n✨ CARACTERÍSTICAS:")
        print(f"  • Bendiciones: {self.stats.get('total_bendiciones', 0)}")
        print(f"  • Maldiciones: {self.stats.get('total_maldiciones', 0)}")
    
    def guardar_resumen_txt(self, filename: str = None) -> str:
        """Guarda el resumen en un archivo de texto."""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(self.dir_salida, f"resumen_{timestamp}.txt")
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("="*60 + "\n")
            f.write("📊 RESUMEN DE SIMULACIÓN\n")
            f.write("="*60 + "\n\n")
            
            f.write(f"📌 DATOS BÁSICOS:\n")
            f.write(f"  • Personajes iniciales: {self.stats.get('total_inicial', 0)}\n")
            f.write(f"  • Personajes vivos: {self.stats.get('vivos', 0)}\n")
            f.write(f"  • Personajes muertos: {self.stats.get('muertos', 0)}\n")
            f.write(f"  • Supervivencia: {self.stats.get('supervivencia', 0):.1f}%\n")
            f.write(f"  • Eventos totales: {self.stats.get('total_eventos', 0)}\n\n")
            
            f.write(f"⚔️ ATRIBUTOS (promedios finales):\n")
            promedios = self.stats.get("promedio_atributos", {})
            for attr in ["F", "RM", "RF", "A", "I", "M", "E", "C"]:
                f.write(f"  • {attr}: {promedios.get(attr, 0):.1f}\n")
            
            f.write(f"\n🏷️ TIPOS DE PERSONAJE:\n")
            tipos = self.stats.get("distribucion_tipos", {})
            total = self.stats.get("vivos", 1)
            for tipo, count in sorted(tipos.items(), key=lambda x: x[1], reverse=True):
                f.write(f"  • {tipo}: {count} ({count/total*100:.1f}%)\n")
            
            f.write(f"\n✨ CARACTERÍSTICAS:\n")
            f.write(f"  • Bendiciones: {self.stats.get('total_bendiciones', 0)}\n")
            f.write(f"  • Maldiciones: {self.stats.get('total_maldiciones', 0)}\n")
        
        print(f"\n📄 Resumen guardado en: {filename}")
        return filename